import pywhatkit as wpp
from openpyxl import load_workbook as xl
import extensions

erro = None

try:
    # abre o arquivo excel
    arquivo_xl = xl("contatosWpp.xlsx")
    aba_contatos = arquivo_xl["Contatos"]
    ultima_linha = aba_contatos.max_row

    lista_celulares = []
    lista_nomes = []

    # verifica os números e nomes da planilha
    for linha in range(2, ultima_linha + 1):
        cell = aba_contatos.cell(row=linha, column=1)
        nome = aba_contatos.cell(row=linha, column=2).value

        if not cell:       
            break

        cell_format = extensions.verificacao(str(cell.value))
        lista_celulares.append(cell_format)
        lista_nomes.append(nome)

except FileNotFoundError:
    erro = FileExistsError
    print("Arquivo Excel não encontrado")


if erro is None:
    try:
        # abre o arquivo.txt para copiar a mensagem de saudação
        with open("saudacao.txt", 'r', encoding="utf-8") as arq_saudacao:
            mensagem = arq_saudacao.read()
            n = 0

        # envia instantaneamente a mensagem
        for cell in lista_celulares:
            print(lista_nomes[n], cell)
            n += 1
            
            wpp.sendwhatmsg_instantly (
                cell,
                mensagem,
                wait_time=10,
                tab_close=True
            )
            print("Mensagem Enviada.\n")
    except FileNotFoundError:
        print("Arquivo de texto não encontrado.")
